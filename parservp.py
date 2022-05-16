import os
import re
import time
from datetime import date, datetime
from threading import Thread
from tqdm.asyncio import tqdm

import openpyxl

# BeautifulSoup
from bs4 import BeautifulSoup

# Other/Common
from fake_useragent import UserAgent

# Selenium
from selenium import webdriver
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EA
from selenium.webdriver.support.ui import WebDriverWait


def driver_input():
    '''Обозначение пути драйвера'''
    try:
        with open('driverpath.txt', 'r', encoding='utf-8') as file:
            src = file.read()
        return src
    except Exception:
        a = input('Путь до драйвера: ')
        with open('driverpath.txt', 'w', encoding='utf-8') as file:
            file.write(a)
        return a


def data_input():
    '''Загрузка входных данных'''
    data = None

    try:
        with open('datapath.txt', 'r', encoding='utf-8') as file:
            src = file.read()
        data = src
    except Exception:
        a = input('Путь до таблицы с входными данными: ')
        with open('datapath.txt', 'w', encoding='utf-8') as file:
            file.write(a)
        data = a

    wb = openpyxl.open(data, read_only=True)
    ws = wb.active

    pro_art = []

    for i, j in ws.rows:
        if ((i.value.upper(), j.value)) not in pro_art:
            pro_art.append((i.value.upper(), j.value))
    print('Входные данные успешно загружены')

    wb.close()
    return pro_art


def get_data_avdmotors(pr_art, driver_path):
    '''Парсинг сайта https://www.avdmotors.ru/'''

    # Настройка драйвера
    user_agent = UserAgent()

    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={user_agent.random}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.headless = True

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)

    # Kоличество всех файлов
    len_url = len(pr_art) - 1

    # Дата сбора информации
    date_1 = date.today()
    now = datetime.now().strftime("%H.%M")
    time_now = date_1.strftime(f'%d.%m.%Y_{now}')

    # Создаем таблицу и удаляем дефолтную запись
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title='data_avdmotors', index=0)
    sheet = wb['data_avdmotors']

    # Задаем названия колонкам в таблице
    sheet.cell(row=1, column=1).value = 'Порядок'
    sheet.cell(row=1, column=2).value = 'Производитель'
    sheet.cell(row=1, column=3).value = 'Артикул'
    sheet.cell(row=1, column=4).value = 'Поставщик'
    sheet.cell(row=1, column=5).value = 'Количество'
    sheet.cell(row=1, column=6).value = 'Срок'
    sheet.cell(row=1, column=7).value = 'Цена'

    count = 2

    # перебор всех артикулов
    for pr, art in tqdm(pr_art, desc='avdmotors', unit_scale=3, unit=' строка', ncols=100):  #unit_scale=3, unit=' строка', ncols=100): bar_format='{desc}{percentage%}'):
        url = f'https://www.avdmotors.ru/price/{art}/{pr}'
        try:
            driver.get(url=url)
            time.sleep(1)

            src = driver.page_source

            soup = BeautifulSoup(src, 'lxml')

            title = soup.find_all('tr')

            len_url -= 1
            count_item = 0
            local_count = 1

            card_list = []

            for item in title:
                try:
                    if count_item == 4:
                        break
                    provider_first = item.find('span').find('b', class_="el-popover__reference").text
                    provider_last = item.find('div', class_="region_2oRao").text
                    provider = f'{provider_first} - {provider_last}'
                    counts = item.find('td', class_="colQuantity_FV_X1").find('span').text
                    term = item.find('td', class_="hidden-sm-and-down colTime__NepP").find('span').text
                    price = item.find('td', class_="hidden-sm-and-down colPrice_3_u-P").find('b', class_="el-tooltip").text
                    price_int = int(price.replace(' ', '')[0:-1])

                    count_item += 1

                    card_list.append([local_count, pr, art, provider, counts, term, price_int])

                except Exception as ex:
                    continue
            result = [card_list[i[1]] for i in sorted([(card_list[0][6], 0), (card_list[1][6], 1), (card_list[2][6], 2), (card_list[3][6], 3)])]
            for j in result[:3]:
                sheet.cell(row=count, column=1).value = local_count
                sheet.cell(row=count, column=2).value = j[1]
                sheet.cell(row=count, column=3).value = j[2]
                sheet.cell(row=count, column=4).value = j[3]
                sheet.cell(row=count, column=5).value = j[4]
                sheet.cell(row=count, column=6).value = j[5]
                sheet.cell(row=count, column=7).value = j[6]

                count += 1
                local_count += 1

            wb.save(f'Результаты сбора данных/avdmotors_{time_now}.xlsx')

        except Exception as e:
            continue

    wb.save(f'Результаты сбора данных/avdmotors_{time_now}.xlsx')
    wb.close()
    print(f'Данные сохранены в таблицу avdmotors_{time_now}.xlsx.')
    driver.quit()


def get_data_tatparts(pr_art, driver_path):
    '''Парсинг сайта https://www.tatparts.ru/'''

    # Настройка драйвера
    user_agent = UserAgent()

    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={user_agent.random}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.headless = True

    dc = DesiredCapabilities.CHROME
    dc['loggingPrefs'] = {'driver': 'OFF', 'server': 'OFF', 'browser': 'OFF'}

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options, desired_capabilities=dc)

    # Kоличество всех файлов
    len_url = len(pr_art) - 1

    # Дата сбора информации
    date_1 = date.today()
    now = datetime.now().strftime("%H.%M")
    time_now = date_1.strftime(f'%d.%m.%Y_{now}')

    # Создаем таблицу и удаляем дефолтную запись
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title='data_tatparts', index=0)
    sheet = wb['data_tatparts']

    # Задаем названия колонкам в таблице
    sheet.cell(row=1, column=1).value = 'Порядок'
    sheet.cell(row=1, column=2).value = 'Производитель'
    sheet.cell(row=1, column=3).value = 'Артикул'
    sheet.cell(row=1, column=4).value = 'Направление'
    sheet.cell(row=1, column=5).value = 'Количество'
    sheet.cell(row=1, column=6).value = 'Срок'
    sheet.cell(row=1, column=7).value = 'Цена'

    count = 2

    print('Данные сайта tatparts обрабатываются...')
    for pr, art in tqdm(pr_art, desc='tatparts', unit_scale=3, unit=' строка', ncols=100):
        url = f'https://www.tatparts.ru/search.html?sort___search_results_by=final_price&article={art}&brand={pr}&_get_block=1'

        # Собираем данные
        try:
            driver.get(url=url)

            src = driver.page_source
            soup = BeautifulSoup(src, 'lxml')

            title = soup.find_all('tr', onmouseout="toggleRow(this)")

            local_count = 1
            len_url -= 1

            card_list = []

            for item in title[0:4]:
                provider = '---' if item.find('td', class_="col_destination_display").text.strip() == '' else item.find(
                    'td', class_="col_destination_display").text.strip()
                counts = 'Есть в наличии' if item.find('td', class_="col_remains").text.strip() == '' else item.find('td', class_="col_remains").text.strip()
                term = item.find('td', class_="col_term").find('div', class_="term_data").text.strip()
                price = item.find('td', class_="col_final_price").find('nobr').text.strip()
                int_price = int(price.replace(' ', '')[0:-2])

                card_list.append([local_count, pr, art, provider, counts, term, int_price])

            result = [card_list[i[1]] for i in
                      sorted([(card_list[0][6], 0), (card_list[1][6], 1), (card_list[2][6], 2), (card_list[3][6], 3)])]

            for j in result[:3]:
                sheet.cell(row=count, column=1).value = local_count
                sheet.cell(row=count, column=2).value = j[1]
                sheet.cell(row=count, column=3).value = j[2]
                sheet.cell(row=count, column=4).value = j[3]
                sheet.cell(row=count, column=5).value = j[4]
                sheet.cell(row=count, column=6).value = j[5]
                sheet.cell(row=count, column=7).value = j[6]

                count += 1
                local_count += 1

                wb.save(f'Результаты сбора данных/tatparts_{time_now}.xlsx')

        except Exception as a:
            continue

    wb.save(f'Результаты сбора данных/tatparts_{time_now}.xlsx')
    wb.close()
    print(f'Данные сохранены в таблицу tatparts_{time_now}.xlsx.')
    driver.quit()


def get_data_trinity_parts(pr_art, driver_path):
    '''Парсинг сайта http://trinity-parts.ru/user/'''

    # Настройка драйвера
    user_agent = UserAgent()

    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={user_agent.random}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.headless = True

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)

    # Kоличество всех файлов
    len_url = len(pr_art) - 1

    # Дата сбора информации
    date_1 = date.today()
    now = datetime.now().strftime("%H.%M")
    time_now = date_1.strftime(f'%d.%m.%Y_{now}')

    url = 'http://trinity-parts.ru/user/'

    print('Аутентификация на сайте trinity_parts началсь...')
    driver.get(url=url)
    time.sleep(3)

    xpath = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/div[1]/input'
    driver.find_element(By.XPATH, xpath).send_keys('LOGIN')
    time.sleep(2)

    xpath_p = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/div[2]/input'
    driver.find_element(By.XPATH, xpath_p).send_keys('PASSWORD')
    time.sleep(2)

    xpath_z = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/label'
    driver.find_element(By.XPATH, xpath_z).click()
    time.sleep(2)

    xpath_s = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/div[4]/div/div[1]/div[2]/div'
    driver.find_element(By.XPATH, xpath_s).click()
    time.sleep(2)

    xpath_n = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/div[4]/div/div[2]/ul/li[5]'
    driver.find_element(By.XPATH, xpath_n).click()
    time.sleep(2)

    xpath_e = '/html/body/div[1]/div/header/div[2]/div/div[2]/div/div[2]/form/div[3]/button'
    driver.find_element(By.XPATH, xpath_e).click()
    time.sleep(2)

    # Создаем таблицу и удаляем дефолтную запись
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title='trinity_parts', index=0)
    sheet = wb['trinity_parts']

    # Задаем названия колонкам в таблице
    sheet.cell(row=1, column=1).value = 'Порядок'
    sheet.cell(row=1, column=2).value = 'Производитель'
    sheet.cell(row=1, column=3).value = 'Артикул'
    sheet.cell(row=1, column=4).value = 'Склад'
    sheet.cell(row=1, column=5).value = 'Количество'
    sheet.cell(row=1, column=6).value = 'Срок'
    sheet.cell(row=1, column=7).value = 'Цена'

    count = 2

    for pr, art in tqdm(pr_art, desc='trinity-parts', unit_scale=3, unit=' строка', ncols=100):
        url = f'http://trinity-parts.ru/user/ajax/opt?#code={art}&producer={pr}'

        # Собираем данные
        try:
            driver.get(url=url)
            driver.refresh()

            wait = WebDriverWait(driver, 10)
            element = wait.until(EA.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div[1]/div/table')))

            src = driver.page_source

            soup = BeautifulSoup(src, 'lxml')

            title4 = soup.find_all('tr', class_=re.compile("^tr-row nophoto requested"))

            len_url -= 1
            local_count = 1

            for item in title4[0:3]:
                provider = item.find('div', class_="tdiv-info").text.strip()
                counts = item.find('td', class_="col-8 t-availability").text.strip()
                term = item.find('td', class_="col-5 t-term").text.strip()[0:3]
                price = item.find('td', class_="col-9 t-price").text.strip()

                sheet.cell(row=count, column=1).value = local_count
                sheet.cell(row=count, column=2).value = pr
                sheet.cell(row=count, column=3).value = art
                sheet.cell(row=count, column=4).value = provider
                sheet.cell(row=count, column=5).value = counts
                sheet.cell(row=count, column=6).value = term
                sheet.cell(row=count, column=7).value = float(price.replace(' ', '').replace(',', '.'))

                count += 1
                local_count += 1

        except Exception as a:
            continue

    wb.save(f'Результаты сбора данных/trinity-parts_{time_now}.xlsx')
    wb.close()
    print(f'Данные сохранены в таблицу trinity-parts_{time_now}.xlsx')
    driver.quit()


def get_data_port3(pr_art, driver_path):
    '''Парсинг сайта https://www.port3.ru'''

    # Настройка драйвера
    user_agent = UserAgent()

    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={user_agent.random}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.headless = True

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)

    # Дата сбора информации
    date_1 = date.today()
    now = datetime.now().strftime("%H.%M")
    time_now = date_1.strftime(f'%d.%m.%Y_{now}')

    # Создаем таблицу и удаляем дефолтную запись
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title='port3', index=0)
    sheet = wb['port3']

    # Задаем названия колонкам в таблице
    sheet.cell(row=1, column=1).value = 'Порядок'
    sheet.cell(row=1, column=2).value = 'Производитель'
    sheet.cell(row=1, column=3).value = 'Артикул'
    sheet.cell(row=1, column=4).value = 'Количество'
    sheet.cell(row=1, column=5).value = 'Срок'
    sheet.cell(row=1, column=6).value = 'Цена'

    count_row = 2

    for pr, art in tqdm(pr_art, desc='port3', unit_scale=3, unit=' строка', ncols=100):
        url = f'https://www.port3.ru/search/{pr}/{art}'
        driver.get(url)
        time.sleep(1)
        driver.refresh()

        src = driver.page_source

        soup = BeautifulSoup(src, 'lxml')

        tbody = soup.find('table', class_="t").find('tbody')
        tr = tbody.find_all('tr')

        count = 0

        local_count = 1

        card_list = []

        for item in tr:
            term = item.find('td', class_="SP dSP").text

            if count == 4:
                break
            counts = 'Много' if item.find('div', class_="avlc").text.split(' ')[0] == '900000' else item.find('div', class_="avlc").text.split(' ')[0]
            price = item.find('td', class_="price").find('div', class_="bold").text

            sheet.cell(row=count_row, column=1).value = local_count
            sheet.cell(row=count_row, column=2).value = pr
            sheet.cell(row=count_row, column=3).value = art
            sheet.cell(row=count_row, column=4).value = counts
            sheet.cell(row=count_row, column=5).value = term
            price_int = int(price.replace(' ', '')[0:-2])
            sheet.cell(row=count_row, column=6).value = price_int

            card_list.append([local_count, pr, art, counts, term, price_int])

            count += 1

        result = [card_list[j[1]] for j in sorted([(i[1][5], i[0]) for i in list(enumerate(card_list))])]


        for j in result[:3]:
            sheet.cell(row=count_row, column=1).value = local_count
            sheet.cell(row=count_row, column=2).value = j[1]
            sheet.cell(row=count_row, column=3).value = j[2]
            sheet.cell(row=count_row, column=4).value = j[3]
            sheet.cell(row=count_row, column=5).value = j[4]
            sheet.cell(row=count_row, column=6).value = j[5]

            count_row += 1
            local_count += 1

            wb.save(f'Результаты сбора данных/port3_{time_now}.xlsx')

        wb.save(f'Результаты сбора данных/port3_{time_now}.xlsx')

    print(f'Данные сохранены в таблицу port3_{time_now}.xlsx')
    driver.quit()
    wb.close()


def main():
    try:
        try:
            os.mkdir("Результаты сбора данных")
        except Exception:
            pass

        pr_art = data_input()
        driver_path = driver_input()

        t1 = Thread(target=get_data_avdmotors, args=(pr_art, driver_path,))
        t2 = Thread(target=get_data_tatparts, args=(pr_art, driver_path,))
        t3 = Thread(target=get_data_trinity_parts, args=(pr_art, driver_path,))
        t4 = Thread(target=get_data_port3, args=(pr_art, driver_path,))

        task = input('''
    --------------------------------------------------------------------------------------
    Номера сайтов:
    1 - avdmotors
    2 - tatparts
    3 - trinity-parts
    4 - port3
    
    Введите слитно номера сайтов данные с которых нужны. Например: 13 (Данные будут собраны с сайтов номер 1 и 3)
    --------------------------------------------------------------------------------------
    Введите данные и нажмите Enter: ''')

        if len(task.strip()) == 1:
            if '1' == task.strip():
                get_data_avdmotors(pr_art, driver_path)

            elif '2' == task.strip():
                get_data_tatparts(pr_art, driver_path)

            elif '3' == task.strip():
                get_data_trinity_parts(pr_art, driver_path)

            elif '4' == task.strip():
                get_data_port3(pr_art, driver_path)

        elif len(task.strip()) == 2:
            if ('1' and '2') in task.strip():
                t1.start()
                t2.start()

                t1.join()
                t2.join()

            elif ('1' and '3') in task.strip():
                t1.start()
                t3.start()

                t1.join()
                t3.join()

            elif ('2' and '3') in task.strip():
                t3.start()
                t2.start()

                t3.join()
                t2.join()

            elif ('1' and '4') in task.strip():
                t1.start()
                t4.start()

                t1.join()
                t4.join()

            elif ('2' and '4') in task.strip():
                t4.start()
                t2.start()

                t4.join()
                t2.join()

            elif ('4' and '3') in task.strip():
                t3.start()
                t4.start()

                t3.join()
                t4.join()

        elif len(task.strip()) == 3:
            if ('1' and '2' and '3') in task.strip():
                t1.start()
                t2.start()
                t3.start()

                t1.join()
                t2.join()
                t3.join()
            elif ('4' and '2' and '3') in task.strip():
                    t4.start()
                    t2.start()
                    t3.start()

                    t4.join()
                    t2.join()
                    t3.join()
            elif ('1' and '4' and '3') in task.strip():
                    t1.start()
                    t4.start()
                    t3.start()

                    t1.join()
                    t4.join()
                    t3.join()
            elif ('1' and '2' and '4') in task.strip():
                    t1.start()
                    t2.start()
                    t4.start()

                    t1.join()
                    t2.join()
                    t4.join()

        else:
            t1.start()
            t2.start()
            t3.start()
            t4.start()

            t1.join()
            t2.join()
            t3.join()
            t4.join()

    except Exception as ex:
        print('Ошибка, проверьте введенные данные. Возможно открыты какие-то таблицы, их нужно закрыть.', ex)
    finally:
        input('\nВсе процессы завершены, для закрытия программы нажмите Enter.')


if __name__ == '__main__':
    main()
